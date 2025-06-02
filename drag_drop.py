import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd


# Path to your master file
MASTER_FILE = r"C:\Users\abdul\Desktop\ALL BARCODES.xlsx"
OUTPUT_FILE = r"C:\Users\abdul\Desktop\NEW_ITEMS_HIGHLIGHTED.xlsx"

# Load all barcodes from 'ALL BARCODES.xlsx' efficiently
def load_master_barcodes():
    try:
        df = pd.read_excel(MASTER_FILE)
        # Convert to string, remove .0 if present, and strip whitespace
        df['BARCODE'] = df['BARCODE'].astype(str).str.split('.').str[0].str.strip()
        #print("[DEBUG] First 10 cleaned barcodes from ALL BARCODES.xlsx:")
        #print(df['BARCODE'].head(10))
        return set(df['BARCODE'])
    except Exception as e:
        print("Error loading master barcodes:", e)
        return set()


# Highlight matching barcodes
def highlight_matching_barcodes(input_file, all_barcodes_set, output_file):
    try:
        wb = load_workbook(input_file)
        ws = wb.active

        # Find "BARCODE" column index
        barcode_col_index = None
        for col_idx, cell in enumerate(ws[1], 1):  # First row (headers)
            if str(cell.value).strip().upper() == "BARCODE":
                barcode_col_index = col_idx
                break

        if not barcode_col_index:
            messagebox.showerror("Error", "Could not find 'BARCODE' column in input file.")
            return False

        # Define styles
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        bold_font = Font(bold=True, color="FFFFFF")  # White text on red background

        match_count = 0

        print(f"[DEBUG] Reading input file '{input_file}'...")
        print(f"[DEBUG] Looking for matches in column '{barcode_col_index}'")

        # Loop through rows and highlight duplicates
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  # Start from second row
            cell = row[barcode_col_index - 1]  # zero-based index
            barcode = str(cell.value).split('.')[0].strip() if cell.value else ""
            if barcode:
                print(f"Row {row_idx}: '{barcode}'")
            if barcode in all_barcodes_set:
                cell.fill = red_fill
                cell.font = bold_font
                match_count += 1

        wb.save(output_file)
        return match_count
    except Exception as e:
        print("Error processing file:", e)
        return False


# Main App Class
class BarcodeHighlighterApp:
    def __init__(self, root):
        self.root = root
        self.all_barcodes_set = load_master_barcodes()

        root.title("📦 Barcode Highlighter Tool")
        root.geometry("500x300")
        root.configure(bg="#f0f0f0")

        self.label = tk.Label(root, text="Select Excel File", bg="#f0f0f0", font=("Arial", 16))
        self.label.pack(pady=30)

        self.status_label = tk.Label(root, text="", bg="#f0f0f0", fg="green", font=("Arial", 12))
        self.status_label.pack(pady=10)

        self.select_button = tk.Button(
            root,
            text="📁 Select File",
            width=30,
            height=2,
            command=self.select_file
        )
        self.select_button.pack(pady=20)

        instructions = tk.Label(
            root,
            text="💡 Click the button above to select an .xlsx file\nThe app will check against ALL BARCODES.xlsx\nand save the highlighted result.",
            bg="#f0f0f0",
            font=("Arial", 10),
            justify="left"
        )
        instructions.pack(pady=10)

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.process_file(file_path)

    def process_file(self, file_path):
        self.status_label.config(text="⏳ Processing file...", fg="orange")
        self.root.update()

        match_count = highlight_matching_barcodes(file_path, self.all_barcodes_set, OUTPUT_FILE)

        if match_count is False:
            self.status_label.config(text="❌ Error occurred during processing.", fg="red")
        elif match_count == 0:
            messagebox.showinfo("Done", "✅ No matching barcodes found.")
            self.status_label.config(text="✅ No matches found.", fg="green")
        else:
            messagebox.showinfo("Done", f"✅ {match_count} matching barcodes highlighted.")
            self.status_label.config(text=f"✅ {match_count} matches found. Output saved.", fg="green")


if __name__ == "__main__":
    root = tk.Tk()
    app = BarcodeHighlighterApp(root)
    root.mainloop()