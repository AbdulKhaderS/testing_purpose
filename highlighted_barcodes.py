from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# File paths
all_barcodes_path = r"C:\Users\abdul\Desktop\ALL BARCODES.xlsx"
new_items_path = r"C:\Users\abdul\Desktop\02-06-25\NEW ITEMS SAM HUSSAIN.xlsx"
output_path = r"C:\Users\abdul\Desktop\NEW_ITEMS_HIGHLIGHTED.xlsx"

# Load all barcodes from 'ALL BARCODES.xlsx' using openpyxl (fast & memory-efficient)
def get_all_barcodes_fast(filepath):
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    barcode_col_index = None

    # Find the "BARCODE" column index
    for row in ws.iter_rows(max_row=1):  # Only first row (headers)
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "BARCODE":
                barcode_col_index = cell.column  # 1-based index
                break

    if barcode_col_index is None:
        print("Error: Could not find 'BARCODE' column in ALL BARCODES.xlsx")
        exit()

    barcodes = set()
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  # Start from second row
        barcode_cell = row[barcode_col_index - 1]  # Convert to 0-based index
        barcode = str(barcode_cell.value).strip() if barcode_cell.value else ""
        if barcode:
            barcodes.add(barcode)

    return barcodes

# Get all barcodes quickly
all_barcodes_set = get_all_barcodes_fast(all_barcodes_path)

# Load new items workbook to apply styling
wb = load_workbook(new_items_path)
ws = wb.active

# Make sure the "BARCODE" column exists in the new file
barcode_col_index = None
for col_idx, cell in enumerate(ws[1], 1):  # First row (headers)
    if cell.value and str(cell.value).strip().upper() == "BARCODE":
        barcode_col_index = col_idx
        break

if barcode_col_index is None:
    print("Error: Could not find 'BARCODE' column in the input Excel file.")
    exit()

# Define styles
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
bold_font = Font(bold=True, color="FFFFFF")  # White bold text on red background

match_count = 0

# Now loop through rows and highlight duplicates
for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  # Start from second row
    cell = row[barcode_col_index - 1]  # zero-based index
    barcode = str(cell.value).strip() if cell.value else ""
    if barcode in all_barcodes_set:
        cell.fill = red_fill
        cell.font = bold_font
        match_count += 1

print(f"\nTotal matching barcodes found and highlighted: {match_count}")
wb.save(output_path)
print(f"Highlighted file saved at: {output_path}")